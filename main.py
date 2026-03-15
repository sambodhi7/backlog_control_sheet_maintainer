from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config import *
from pprint import pprint
from shortFormData import shortFormData

thin_side = Side(border_style="thin", color="000000")
table_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
header_font = Font(name='Calibri', size=11, bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Light Grey
body_font = Font(name='Calibri', size=11)
center_aligned = Alignment(horizontal='center', vertical='center')
left_aligned = Alignment(horizontal='left', vertical='center')

def transformShortFormData():
    lookup = dict()
    for key, values in shortFormData.items():
        for v in values :
            lookup[v] = key 
    return lookup

codeLookup = transformShortFormData()

def apply_formatting(cell, is_header=False, align='center'):
    cell.border = table_border
    if is_header:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
    else:
        cell.font = body_font
        cell.alignment = left_aligned if align == 'left' else center_aligned

def get_control_dict(sheet):
    r = config.ROW_STARTING 
    control_dict = {}
    while sheet.cell(row=r, column=1).value is not None:
        name = sheet.cell(row=r, column=3).value
        btid = sheet.cell(row=r, column=2).value
        toenter = 4 
        subject_sets = set()
        while(sheet.cell(row=r, column=toenter).value is not None):    
            subject =  (sheet.cell(row=r,column=toenter).value).strip()
            if subject in codeLookup:
                subject = codeLookup[subject]
            subject_sets.add(subject)
            toenter+=1
        clear_size = toenter - 4
        
        toenter=4
        control_dict[btid] = {
            'row' : r,
            'name' : name,
            'toenter': toenter,
            'subjects_set': subject_sets,
            'clear_size': clear_size
        }
        r+=1
    control_dict["last_row"] = r
    return control_dict

def process_subject_sheet(sheet, control, newdict, namelookup):
    def find_course_name():
        c=1
        r=config.COURSE_PAGE_HEADER_WITH_COURSE_TITLE_ROW_NO-1
        while sheet.cell(row=r, column=c).value != 'Cource Title':
            c+=1
        name = sheet.cell(row=r+1,column=c).value
        return str(name).strip().replace('\n', ' ') if name else "Unknown Course"
    
    def get_cols():
        r = config.COURSE_PAGE_HEADER_WITH_COURSE_TITLE_ROW_NO + 1
        c = 1
        res = {'reexam': -1, 'name': -1, 'btid': -1}
        while True:
            v = sheet.cell(row=r, column=c).value
            if v:
                vl = str(v).strip().replace('\n', ' ')
                if vl == "Re-Exam Grades": res['reexam'] = c; break
                elif vl == "Name": res['name'] = c
                elif vl == "Roll No.": res['btid'] = c
            if c >= 1000: raise Exception("Header not found")
            c += 1
        return res['reexam'], res['name'], res['btid']

    course_title = find_course_name()
    if course_title in codeLookup:
        course_title = codeLookup[course_title]
    reexam_grade_col, name_col, btid_col = get_cols()
    r = config.COURSE_PAGE_HEADER_WITH_COURSE_TITLE_ROW_NO + 1
    while sheet.cell(row=r, column=1).value != 1: r += 1

    while True:
        n_v, b_v, g_v = [sheet.cell(row=r, column=col).value for col in [name_col, btid_col, reexam_grade_col]]
        if n_v is None and b_v is None: break
        
        name, btid, grade = [str(v).strip() if v else "" for v in [n_v, b_v, g_v]]
        namelookup[btid] = name
        if btid in control:
            newdict[btid]= list(control[btid].get("subjects_set"))
        if grade == "FF":
            if btid in control:
                if course_title not in control[btid].get('subjects_set', set()):
                    newdict.setdefault(btid, []).append(course_title)
            else:
                newdict.setdefault(btid, []).append(course_title)
        else:
            if btid in control :
                if course_title in control[btid].get('subjects_set', set()):
                    newdict[btid].remove(course_title)
                    control[btid]['subjects_set'].remove(course_title)
        r += 1

def process_subject_file(file, control, newdict, namelookup):
    subject_wb = load_workbook(file, data_only=True)
    for sn in subject_wb.sheetnames:
        process_subject_sheet(subject_wb[sn], control, newdict, namelookup)
   
def save_to_control_file(sheet, control, newdict, namelookup):
    row_h = 25 
   
    print("Newdict:")
    for btid, subs in newdict.items():
        print(f"{btid}: {subs}")
   
   
    for btid in newdict:
        
        if btid in control:
            r = control[btid]['row']
            toenter = control[btid]['toenter']
            clear_size = control[btid]['clear_size']
            for c in range(4, 4 + clear_size):
                sheet.cell(row=r, column=c).value = None
            for sub in newdict[btid]:
                if sub in shortFormData:
                    sub = shortFormData[sub][0]
                
                sheet.cell(row=r, column=toenter).value = sub
                toenter += 1
        else:
            r = control['last_row']
            sheet.cell(row=r, column=1).value = r - config.ROW_STARTING + 1
            sheet.cell(row=r, column=2).value = btid
            sheet.cell(row=r, column=3).value = namelookup.get(btid, "Name Unknown")
            toenter = 4
            
            for sub in newdict[btid]:
                if sub in shortFormData:
                    sub = shortFormData[sub][0]
                sheet.cell(row=r, column=toenter).value = sub
               
                toenter += 1
            control['last_row'] += 1

  
    for row in range(config.ROW_STARTING, sheet.max_row + 1):
        sheet.row_dimensions[row].height = row_h
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
          
            align = 'left' if col == 3 else 'center'
            apply_formatting(cell, align=align)


    for col in range(1, sheet.max_column + 1):
        max_l = 0
        column_letter = get_column_letter(col)
        for cell in sheet[column_letter]:
            if cell.value: max_l = max(max_l, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_l + 4

    # sheet.parent.save("control_updated.xlsx")

    # delete all the rows in the control file for which there are no entries in the newdict
    # and shift up the rows below it
    rows_to_delete = []
    for btid, data in control.items():
        if btid != "last_row" and (btid not in newdict or len(newdict[btid]) == 0):
            rows_to_delete.append(data['row'])
    rows_to_delete.sort(reverse=True)
    for row in rows_to_delete:
        sheet.delete_rows(row, 1)
    
    # Renumber the serial numbers in column 1
    for row in range(config.ROW_STARTING, sheet.max_row + 1):
        sheet.cell(row=row, column=1).value = row - config.ROW_STARTING + 1
    
    sheet.parent.save("control_updated.xlsx") 


def main():
    control_wb = load_workbook("control.xlsx")
    control_sheet = control_wb.active
    c_dict = get_control_dict(control_sheet)
    n_dict, n_lookup = {}, {}
    
    process_subject_file("csh2.xlsx", c_dict, n_dict, n_lookup)
    process_subject_file("csh4.xlsx", c_dict, n_dict, n_lookup)
    
    save_to_control_file(control_sheet, c_dict, n_dict, n_lookup)

if __name__ == "__main__":
    main()
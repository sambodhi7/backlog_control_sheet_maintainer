from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from config import config
from pprint import pprint
import json
shortFormData = json.load(open("shortFormData.json", "r", encoding="utf-8"))


def util_format_string(s):
    if s is None:
        return ""
    return s.strip().upper().replace(" ", "")



def transformShortFormData():
    lookup = dict()
    for key, values in shortFormData.items():
        for v in values :
            v = util_format_string(v)
            lookup[v] = key 
    return lookup

codeLookup = transformShortFormData()

btid_to_name_lookup = dict()


def get_control_dict(sheet):
    r = config.ROW_STARTING 
    control_dict = {}
    while sheet.cell(row=r, column=1).value is not None:
        name = sheet.cell(row=r, column=3).value
        btid = sheet.cell(row=r, column=2).value
        toenter = 4 
        subject_sets = set()
        while(sheet.cell(row=r, column=toenter).value is not None):    
            subject =  (sheet.cell(row=r,column=toenter).value).strip()
            if subject in codeLookup:
                subject = codeLookup[subject]
            subject_sets.add(subject)
            toenter+=1
        clear_size = toenter - 4
        
        toenter=4
        control_dict[btid] = {
            'row' : r,
            'name' : name,
            'toenter': toenter,
            'subjects_set': subject_sets,
            'clear_size': clear_size
        }
        r+=1
    control_dict["last_row"] = r
    return control_dict





def process_s_term_grade_val(s):
    s=s.strip()
    s= s.replace("," , " ")
    s = s.replace("S.TERM", "")
    for part in s.split(")"):
        part = part.strip()
        if len(part)==0:
            continue
        grade, p2 = part.split("(")

        p2 = p2.replace(","," ")
        sub=p2.split("CR")[0]
        sub=sub.strip()
        yield grade.strip(), sub.strip()

def transform_codelookup_to_shortform(codeLookup):
    data =shortFormData
    for title, code in codeLookup.items():
     
        if code in data:
            if title not in data[code]:
                data[code].append(title)
        else:
            data[code] = [title]
   
    json.dump(data, open("shortFormData.json", "w", encoding="utf-8"), indent=4, ensure_ascii=False)
            

temp = [] 
def process_subject_sheet2(sheet, control, newdict, codeLookup):
    row = config.HEADERR_STARTING
    c = 1
    cell = sheet.cell(row = row ,column = c)
    while util_format_string(cell.value)!='COURSECODE':
        c+=1
        cell = sheet.cell(row = row ,column = c)
        if c > 20:
            print("Course code header not found in sheet", sheet.title)
            return
    c+=1

    # print(sheet.cell(row = row ,column = c).value)

    def get_row_to_btid_dict(): 
        r = config.ROW_STARTING
        res = {}
        while True:
            value  = sheet.cell(row=r, column=1).value
            value = util_format_string(value)
            if "NO" in value:
                break 
            r+=1
        r+=1
        rs=r
        
        while True:
            value  = sheet.cell(row=r, column=1).value
            if value is None:
                break 
            btid = sheet.cell(row=r, column=2).value
            name = sheet.cell(row=r, column=3).value
            if btid is not None:
                res[r] = util_format_string(btid)
                btid_to_name_lookup[btid] = name
    
            r+=1
        return rs, res
    
    row_s, row_to_btid_dict = get_row_to_btid_dict()
    
    col_to_course_code = dict() 
    while(1):
        course_code_cell = sheet.cell(row = row ,column = c)
        if course_code_cell.value is None:
            break
        
        grade_cell = sheet.cell(row = row_s ,column = c).value
        ct = 0
        course_code = util_format_string(course_code_cell.value)
        course_title = sheet.cell(row=row+1 ,column = c).value
        if course_code not in codeLookup:
            codeLookup[course_title] = course_code
        # while(ct<config.MAX_NUMBER_OF_STUDENTS): 
        #     grade_cell = sheet.cell(row = row_s + ct ,column = c).value
        #     if row_to_btid_dict.get(row_s + ct) is None:
        #         break
        #     if grade_cell is None:
        #         ct+=1
        #         continue
        #     grade = util_format_string(grade_cell.value)
        #     btid = row_to_btid_dict[row_s + ct]
            
        #     course_title = sheet.cell(row=row+1 ,column = c).value
        #     course_code = util_format_string(course_code_cell.value)
            
        #     if course_code not in codeLookup:
        #         codeLookup[course_code] = course_title
                        
        #     ct+=1
        col_to_course_code[c] = course_code

        c+=1
    
    def process_course_and_grade(btid, course_code, grade):
                prev = course_code
                
                if course_code in codeLookup:
                    course_code = codeLookup[course_code]
                else:
                    temp.append(course_code)
                if btid in control and course_code in control[btid]['subjects_set']:
                    if "FF" not in grade: 
                        newdict[btid]['removed'].append(course_code)
                        if course_code in newdict[btid]['list']:
                            newdict[btid]['list'].remove(course_code)

                if btid in newdict and course_code in newdict[btid]['list']:
                    if "FF" not in grade:
                        newdict[btid]['list'].remove(course_code)
                        if course_code in newdict[btid]['added']:
                            newdict[btid]['added'].remove(course_code)
                        

                if "FF" in grade:
                    if btid in newdict:
                        if course_code not in newdict[btid]['list']:
                            newdict[btid]['added'].append(course_code)
                            newdict[btid]['list'].append(course_code)
                    else:
                        newdict[btid] = {
                            "list" : [course_code],
                            "added" : [course_code],
                            "removed" : []
                        }
            

    for r, btid in row_to_btid_dict.items():
        maxc = 0 
        for c, course_code in col_to_course_code.items():
            maxc = max(maxc, c)
            grade_cell = sheet.cell(row = r ,column = c).value
            if grade_cell is None:
                continue
            grade = util_format_string(grade_cell)
            if grade == "":
                continue
            course_code_cell = sheet.cell(row=row ,column = c)
            course_code = util_format_string(course_code_cell.value)
            # if course_code not in codeLookup:
            #     course_title = sheet.cell(row=row+2 ,column = c).value
            #     codeLookup[course_code] = course_title
            process_course_and_grade(btid, course_code, grade)
        
        #find the next column with a ( in the valuue means we will calculate backlog thing too for summer terms uk
        next_c = maxc + 1
        while True:
            cell = sheet.cell(row = r ,column = next_c)
            if cell.value is None:
                break
            val = f"{cell.value}"
            val = util_format_string(val)
            if "(" in val and ")" in val:
                for grade, course_code in process_s_term_grade_val(val):
                    process_course_and_grade(btid, course_code, grade)
            next_c+=1
    
    
            

def make_new_dict_from_control(control):
    newdict = {}

    for btid in control:
        newdict[btid] = {
            "list" : list(control[btid].get("subjects_set", set())),
            "added" : [],
            "removed" : []
        } 
        btid_to_name_lookup[btid] = control[btid]['name']
    return newdict

def process_subject_file2(file, control, newdict, codeLookup):
    subject_wb = load_workbook(file, data_only=True)

    for sn in subject_wb.sheetnames:
        sn = sn.upper()
        if "SEM" in sn:
            process_subject_sheet2(subject_wb[sn], control, newdict, codeLookup)




# make the save to file function it will just give a file name and you will genrate the file it should have the slno, bitid, name, and the the subjects and in in another sheet for every btid show the subjects added and deleted, the added ones are in green and the deleted ones in red 
def save_to_control_file2(output_path, control, newdict, namelookup, last_row):
    
    
    def get_subject_name(subject_code):
        if subject_code in shortFormData:
            return shortFormData[subject_code][0]
    
    control_wb = Workbook()
    sheet = control_wb.active
    sheet.title = 'Control'
    # Populate the control sheet
    sheet['A1'] = 'Sl.No'
    sheet['B1'] = 'BTID'
    sheet['C1'] = 'Name'
    
    row_num = 2
    slno = 1
    for btid in newdict.keys():
        if newdict[btid]['list'] == []:
            continue
        sheet.cell(row=row_num, column=1).value = slno
        sheet.cell(row=row_num, column=2).value = btid
        sheet.cell(row=row_num, column=3).value = btid_to_name_lookup.get(btid, "")
        
        subjects = newdict[btid]['list']
        col = 4
        for subj in subjects:
            
            sheet.cell(row=row_num, column=col).value = get_subject_name(subj) or subj
            
            col += 1
        
        row_num += 1
        slno += 1
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H.%M.%S")
    changes_sheet = control_wb.create_sheet(f"changes {timestamp}")
    changes_sheet['A1'] = 'BTID'

    max_added = max((len(newdict[btid]['added']) for btid in newdict), default=0)
    max_removed = max((len(newdict[btid]['removed']) for btid in newdict), default=0)

    for i in range(max_added):
        changes_sheet.cell(row=1, column=2 + i).value = f'Added {i+1}'
    for i in range(max_removed):
        changes_sheet.cell(row=1, column=2 + max_added + i).value = f'Removed {i+1}'

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    row_num = 2
    for btid in sorted(newdict.keys()):
        if newdict[btid]['list'] == []:
            continue
        changes_sheet.cell(row=row_num, column=1).value = btid

        for i, subj in enumerate(sorted(newdict[btid]['added'])):
            cell = changes_sheet.cell(row=row_num, column=2 + i)
            cell.value = get_subject_name(subj) or subj
            cell.fill = green_fill

        for i, subj in enumerate(sorted(newdict[btid]['removed'])):
            cell = changes_sheet.cell(row=row_num, column=2 + max_added + i)
            cell.value = get_subject_name(subj) or subj
            cell.fill = red_fill

        row_num += 1
    transform_codelookup_to_shortform(codeLookup)
    control_wb.save(output_path)
    tempx = set(temp)
    tempx = list(tempx)
    for x in tempx : 
        if x not in codeLookup.values():
            print(x)
    



def main():
    # control_dict = get_control_dict(load_workbook("control2.xlsx").active)
    
    # last_row = control_dict.get("last_row", 1000)
    # control_dict.pop("last_row", None)
    # newdict = make_new_dict_from_control(control_dict)
    # process_subject_file2("newdata.xlsx", control_dict, newdict, codeLookup)
    # # make a new file to save
    # save_to_control_file2("updated_control2.xlsx", control_dict, newdict, {}, last_row)
    c = get_control_dict(load_workbook("updated_control2.xlsx").active)
    
if __name__ == "__main__":
    main()